import requests
from bs4 import BeautifulSoup

import openpyxl

#엑셀파일 생성
xlFile = openpyxl.Workbook()
#워크시트 만들기
ws = xlFile.create_sheet('주식현재가')
#현재 활성화된 시트를 선택하려면 xlFile.active()이용

#데이터입력
ws['A1'] = '종목코드'
ws['B1'] = '현재가'

#종목 코드 리스트(여러 종목을 가져오기 위해)
codes = [
    '005930', #삼성전자
    '000660', #sk하이닉스
    '035720' #카카오
]

row = 2
 
for code in codes :
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one("#_nowVal").text 
    #그냥 출력하면 75,000처럼 ,가 있는 상태로 출력되고 문자열이기 때문에 아래처럼 대체해줌
    price = price.replace(',', '')
    
    ws[f'A{row}'] = code
    ws[f'B{row}'] = int(price)
    row = row+1


#엑셀저장
xlFile.save(r'D:\Crawling_Python\01_네이버_주식현재가_크롤링\assignment_현재가_저장.xlsx')

