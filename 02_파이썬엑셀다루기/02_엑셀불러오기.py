import openpyxl

#경로가 반복되는 것을 막기 위해 변수에 저장해줌
fpath = r'D:\Crawling_Python\02_파이썬엑셀다루기\참가자_data.xlsx'
#1. 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

#2. 엑셀 시트 선택
ws = wb['오징어게임']

#3. 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'

#4. 엑셀 저장
wb.save(fpath)